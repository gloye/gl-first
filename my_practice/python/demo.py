'''
# demo
'''
import json
import re
import requests
import bs4
import openpyxl


URL = "https://www.billboard.com/charts/hot-100"


print('Downloading page %s...' % URL)
RES = requests.get(URL)
RES.raise_for_status()

SOUP = bs4.BeautifulSoup(RES.text, 'html5lib')

LIST = SOUP.select('.chart-row')
LIST_LEN = len(LIST)

WB = openpyxl.Workbook()
SHEET = WB.active
SHEET.title = '榜单'
SHEET['A1'] = '排名'
SHEET['B1'] = '歌名'
SHEET['C1'] = '歌手'
SHEET['D1'] = '链接'


def parse_jsonp(jsonp_str):
    '''
    #  解析JSONP
    '''
    try:
        return re.search('^[^(]*?\((.*)\)[^)]*$', jsonp_str).group(1)
    except:
        raise ValueError('Invalid JSONP')


def getLink(keyword):
    '''
    # 获取连接
    '''
    search_url = 'http://soapi.yinyuetai.com/search/video-search?callback=jQuery110207370039710434122_1516538494517&keyword=%s&pageIndex=1&pageSize=24&offset=0&orderType=&area=&property=&durationStart=0&durationEnd=&regdateStart=&regdateEnd=1516538495&clarityGrade=&source=&thirdSource=&_=1516538494534' % keyword
    search_raw = requests.get(search_url)
    search_raw.raise_for_status()
    data = json.loads(parse_jsonp(search_raw.text))
    if data['videos']['data']:
        video_id = data['videos']['data'][0]['id']
    else:
        return '搜索不到'
    video_url = 'http://v.yinyuetai.com/video/%s' % video_id
    return video_url


for i in range(LIST_LEN):
    row = i + 2
    rank = LIST[i].select('.chart-row__current-week')[0].getText()
    song = LIST[i].select('.chart-row__song')[0].getText()
    artist = LIST[i].select('.chart-row__artist')[0].getText()
    keyword = song + ' ' + artist
    link = getLink(keyword)
    SHEET.cell(row=row, column=1).value = rank
    SHEET.cell(row=row, column=2).value = song
    SHEET.cell(row=row, column=3).value = artist
    SHEET.cell(row=row, column=4).value = link
    print('%s 已完成' % i)

WB.save('list.xlsx')
